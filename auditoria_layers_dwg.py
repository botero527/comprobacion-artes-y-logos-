#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
AUDITORÍA LAYERS K / K2 — AGP PLANOS TECNICOS  (AutoCAD COM)
Estructura: RUTA_BASE > Vehículo > Modelo > Versión > ARTES
Recolecta:  DWGs sueltos en ARTES  +  subcarpetas con "BN" en el nombre
Excluye:    subcarpetas con "OBSOLETO" en el nombre
Estado:
  ACTUALIZADA  → Layer K existe y es AZUL (ACI 5)  +  Layer K2 existe y es VERDE (ACI 3)
  VIEJA        → No hay Layer K ni Layer K2
  INCOMPLETA   → Existe alguno pero falta color correcto o falta el otro
  ERROR        → Archivo no se pudo abrir

SISTEMA DE CHECKPOINT:
  - Guarda el progreso en un archivo JSON después de cada vehículo
  - Si AutoCAD se cae, al volver a ejecutar retoma desde donde quedó
  - El Excel se puede regenerar en cualquier momento desde el checkpoint
  - Para reiniciar desde cero: borra el archivo .json o usa --reiniciar
"""

import os
import sys
import json
import time
import threading
import argparse
from datetime import datetime

# ──────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────
RUTA_BASE        = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
ARCHIVO_EXCEL    = f"Auditoria_Layers_K_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
ARCHIVO_CHECKPOINT = "auditoria_checkpoint.json"   # guarda el progreso
COLOR_AZUL_ACI   = 5
COLOR_VERDE_ACI  = 3
# ──────────────────────────────────────────────────────────

try:
    import win32com.client
    import pythoncom
except ImportError:
    print("Falta pywin32. Ejecuta: pip install pywin32")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────
# LOGGER
# ──────────────────────────────────────────────────────────
class Logger:
    def _ts(self): return time.strftime("%H:%M:%S")
    def info(self, m):  print(f"{self._ts()}  {m}")
    def warn(self, m):  print(f"{self._ts()}  [!] {m}")
    def error(self, m): print(f"{self._ts()}  [X] {m}")
    def ok(self, m):    print(f"{self._ts()}  [✓] {m}")

log = Logger()


# ──────────────────────────────────────────────────────────
# CHECKPOINT — guardar y cargar progreso
# Estructura JSON:
#   {
#     "completados": { vehiculo: [filas], ... },   ← vehículos 100% listos
#     "en_curso":    { "vehiculo": "CHEVROLET",    ← vehículo parcialmente procesado
#                      "filas":   [filas...] }
#   }
# ──────────────────────────────────────────────────────────
GUARDAR_CADA_N = 5   # guardar checkpoint cada N archivos dentro de un vehículo

def checkpoint_cargar(ruta_archivo):
    if not os.path.exists(ruta_archivo):
        return {}, None, []
    try:
        with open(ruta_archivo, "r", encoding="utf-8") as f:
            raw = json.load(f)
        completados = raw.get("completados", {})
        en_curso    = raw.get("en_curso", {})
        vehiculo_parcial = en_curso.get("vehiculo")
        filas_parciales  = en_curso.get("filas", [])
        total_dwg = sum(len(v) for v in completados.values()) + len(filas_parciales)
        log.ok(f"Checkpoint: {len(completados)} vehículo(s) completos, "
               f"{len(filas_parciales)} DWG(s) del vehículo en curso "
               f"({vehiculo_parcial or '—'})  —  {total_dwg} DWG(s) en total")
        return completados, vehiculo_parcial, filas_parciales
    except Exception as e:
        log.warn(f"No se pudo leer el checkpoint ({e}), se empieza desde cero")
        return {}, None, []


def checkpoint_guardar(ruta_archivo, completados, vehiculo_en_curso=None, filas_en_curso=None):
    try:
        payload = {
            "completados": completados,
            "en_curso": {
                "vehiculo": vehiculo_en_curso or "",
                "filas":    filas_en_curso   or [],
            }
        }
        tmp = ruta_archivo + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        os.replace(tmp, ruta_archivo)
    except Exception as e:
        log.warn(f"Error guardando checkpoint: {e}")


def checkpoint_borrar(ruta_archivo):
    for f in [ruta_archivo, ruta_archivo + ".tmp"]:
        try:
            os.remove(f)
        except FileNotFoundError:
            pass


# ──────────────────────────────────────────────────────────
# COLORES ACI
# ──────────────────────────────────────────────────────────
_COLORES = {
    1: "Rojo", 2: "Amarillo", 3: "Verde", 4: "Cyan",
    5: "Azul", 6: "Magenta",  7: "Blanco/Negro",
    8: "Gris oscuro", 9: "Gris claro",
    0: "ByBlock", 256: "ByLayer",
}
def nombre_color(aci): return _COLORES.get(abs(aci), f"ACI {abs(aci)}")


# ──────────────────────────────────────────────────────────
# MOTOR AUTOCAD
# ──────────────────────────────────────────────────────────
class AutoCAD:
    def __init__(self):
        pythoncom.CoInitialize()
        try:
            self.app = win32com.client.GetActiveObject("AutoCAD.Application")
            try:
                self.app.Preferences.OpenSave.DemandLoadARXApp = 2
            except Exception:
                pass
            log.ok(f"AutoCAD conectado: {self.app.Version}")
        except Exception as e:
            log.error(f"No hay AutoCAD abierto: {e}")
            log.error("Abre AutoCAD (sin ningún archivo) y vuelve a ejecutar.")
            sys.exit(1)

    def _suprimir_dialogs(self):
        """
        Desactiva todo lo que puede hacer que AutoCAD se trabe al abrir un DWG:
          XLOADCTL 0  → no carga XREFs (causa #1 de trabas en red)
          FILEDIA  0  → suprime diálogos de archivos
          EXPERT   5  → suprime todos los prompts de confirmación
          PROXYSHOW 0 → no intenta dibujar objetos proxy
        """
        for var, val in [("XLOADCTL", 0), ("FILEDIA", 0), ("EXPERT", 5), ("PROXYSHOW", 0)]:
            try:
                self.app.SetSystemVariable(var, val)
            except Exception:
                pass

    def _restaurar_dialogs(self):
        """Restaura variables para uso normal de AutoCAD."""
        for var, val in [("XLOADCTL", 2), ("FILEDIA", 1), ("EXPERT", 0), ("PROXYSHOW", 1)]:
            try:
                self.app.SetSystemVariable(var, val)
            except Exception:
                pass

    def vivo(self):
        """Verifica que AutoCAD sigue respondiendo. Intenta varias veces antes de declarar caída."""
        for _ in range(3):
            try:
                _ = self.app.Version
                return True
            except Exception:
                time.sleep(1.0)
        return False

    def leer_layers_con_timeout(self, ruta, timeout=25):
        """
        Abre el DWG, lee los layers y cierra — todo en un hilo secundario.
        Si tarda más de `timeout` segundos, lo cancela y retorna None.
        Retorna: (dict_layers, None) si ok, (None, msg_error) si falla.
        """
        ruta_abs = os.path.abspath(ruta)
        self._suprimir_dialogs()

        layers_result = [None]
        error_result  = [None]

        try:
            stream = pythoncom.CoMarshalInterThreadInterfaceInStream(
                pythoncom.IID_IDispatch, self.app
            )
        except Exception as e:
            error_result[0] = f"Marshal falló: {e}"
            return None, error_result[0]

        def _worker():
            pythoncom.CoInitialize()
            doc = None
            try:
                app_hilo = win32com.client.Dispatch(
                    pythoncom.CoGetInterfaceAndReleaseStream(
                        stream, pythoncom.IID_IDispatch
                    )
                )
                doc = app_hilo.Documents.Open(ruta_abs, True)
                time.sleep(0.4)

                layers = {}
                col = doc.Layers
                for i in range(col.Count):
                    try:
                        l   = col.Item(i)
                        aci = abs(l.Color)
                        layers[l.Name] = {
                            "color_aci":   aci,
                            "color_texto": nombre_color(aci),
                        }
                    except Exception:
                        continue
                layers_result[0] = layers

            except Exception as e:
                error_result[0] = str(e)[:80]
            finally:
                if doc is not None:
                    try:
                        doc.Close(False)
                        time.sleep(0.1)
                    except Exception:
                        pass
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        t.join(timeout)

        if t.is_alive():
            log.warn(f"  TIMEOUT ({timeout}s) — saltando archivo colgado")
            # Intentar cerrar el documento trabado desde el hilo principal
            try:
                for i in range(self.app.Documents.Count - 1, -1, -1):
                    try:
                        doc = self.app.Documents.Item(i)
                        if "Drawing1" not in doc.Name:
                            doc.Close(False)
                            time.sleep(1.5)
                            break
                    except Exception:
                        pass
            except Exception:
                pass
            return None, "TIMEOUT — archivo colgado saltado"

        if error_result[0]:
            return None, error_result[0]

        return layers_result[0], None

    def cerrar_docs_abiertos(self):
        """Cierra cualquier documento que haya quedado abierto (excepto Drawing1)."""
        try:
            cerrados = 0
            for i in range(self.app.Documents.Count - 1, -1, -1):
                try:
                    doc = self.app.Documents.Item(i)
                    if "Drawing1" not in doc.Name:
                        doc.Close(False)
                        time.sleep(0.1)
                        cerrados += 1
                except Exception:
                    pass
            return cerrados
        except Exception:
            return 0

    def quit(self):
        self._restaurar_dialogs()   # dejar AutoCAD usable después del proceso
        self.cerrar_docs_abiertos()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ──────────────────────────────────────────────────────────
# LÓGICA DE ESTADO
# ──────────────────────────────────────────────────────────
def evaluar_estado(layers_dict):
    nombre_k = nombre_k2 = color_k = color_k2 = ""
    tiene_k = tiene_k2 = k_azul = k2_verde = False

    for nombre, info in layers_dict.items():
        u = nombre.upper().strip()
        if u == "K":
            tiene_k  = True;  nombre_k = nombre
            color_k  = info["color_texto"]
            k_azul   = info["color_aci"] == COLOR_AZUL_ACI
        if u == "K2":
            tiene_k2  = True; nombre_k2 = nombre
            color_k2  = info["color_texto"]
            k2_verde  = info["color_aci"] == COLOR_VERDE_ACI
            
    if k_azul and k2_verde:
        estado = "ACTUALIZADA"
    elif not tiene_k and not tiene_k2:
        estado = "VIEJA"
    else:
        estado = "INCOMPLETA"
    return {
        "estado":    estado,
        "tiene_k":   tiene_k,  "nombre_k":  nombre_k,  "color_k":  color_k,  "k_azul":   k_azul,
        "tiene_k2":  tiene_k2, "nombre_k2": nombre_k2, "color_k2": color_k2, "k2_verde": k2_verde,
    }


# ──────────────────────────────────────────────────────────
# RECOLECCIÓN DE DWGs
# ──────────────────────────────────────────────────────────
def recolectar_dwgs(ruta_artes):
    dwgs = []
    try:
        items = os.listdir(ruta_artes)
    except Exception as e:
        log.warn(f"No se puede leer {ruta_artes}: {e}")
        return dwgs

    for item in items:
        ruta_item = os.path.join(ruta_artes, item)
        if os.path.isfile(ruta_item) and item.lower().endswith(".dwg"):
            dwgs.append(("suelto", ruta_item))
        elif os.path.isdir(ruta_item) and "BN" in item.upper():
            if "OBSOLETO" in item.upper():
                continue
            try:
                for sub in os.listdir(ruta_item):
                    ruta_sub = os.path.join(ruta_item, sub)
                    if os.path.isdir(ruta_sub) and "OBSOLETO" in sub.upper():
                        continue
                    if os.path.isfile(ruta_sub) and sub.lower().endswith(".dwg"):
                        dwgs.append((f"BN/{item}", ruta_sub))
            except Exception as e:
                log.warn(f"Error leyendo carpeta BN {item}: {e}")

    return sorted(set(dwgs), key=lambda x: x[1])


# ──────────────────────────────────────────────────────────
# PROCESAR UN VEHÍCULO COMPLETO
# ──────────────────────────────────────────────────────────
def procesar_vehiculo(vehiculo, ruta_vehiculo, motor, completados,
                      filas_previas=None, rutas_ya_procesadas=None):
    """
    Procesa todos los DWGs de un vehículo.
    - filas_previas: filas ya guardadas en checkpoint para este vehículo (retoma)
    - rutas_ya_procesadas: set de rutas ya hechas para no repetirlas
    Guarda checkpoint cada GUARDAR_CADA_N archivos.
    Retorna lista de filas completas, o None si AutoCAD se cayó.
    """
    filas                = list(filas_previas or [])
    rutas_ya_procesadas  = set(rutas_ya_procesadas or [f["ruta"] for f in filas])
    nuevos_desde_guardado = 0

    modelos = sorted(d for d in os.listdir(ruta_vehiculo)
                     if os.path.isdir(os.path.join(ruta_vehiculo, d)))

    for modelo in modelos:
        ruta_modelo = os.path.join(ruta_vehiculo, modelo)
        versiones = sorted(d for d in os.listdir(ruta_modelo)
                           if os.path.isdir(os.path.join(ruta_modelo, d)))

        for version in versiones:
            ruta_version = os.path.join(ruta_modelo, version)

            ruta_artes = None
            try:
                for sub in os.listdir(ruta_version):
                    if sub.upper() == "ARTES" and \
                       os.path.isdir(os.path.join(ruta_version, sub)):
                        ruta_artes = os.path.join(ruta_version, sub)
                        break
            except Exception:
                continue

            if not ruta_artes:
                continue

            dwgs = recolectar_dwgs(ruta_artes)
            if not dwgs:
                continue

            pendientes_version = [(o, p) for o, p in dwgs if p not in rutas_ya_procesadas]
            saltados_version   = len(dwgs) - len(pendientes_version)

            if saltados_version:
                log.info(f"    {modelo} / {version}: {len(dwgs)} DWG(s)  "
                         f"({saltados_version} ya procesados, {len(pendientes_version)} pendientes)")
            else:
                log.info(f"    {modelo} / {version}: {len(dwgs)} DWG(s)")

            for origen, dwg_path in pendientes_version:
                nombre = os.path.basename(dwg_path)
                fila = {
                    "vehiculo": vehiculo, "modelo": modelo, "version": version,
                    "archivo":  nombre,   "origen": origen, "ruta":    dwg_path,
                    "estado":        "ERROR",
                    "tiene_k":       False, "nombre_k":  "", "color_k":  "", "k_azul":   False,
                    "tiene_k2":      False, "nombre_k2": "", "color_k2": "", "k2_verde": False,
                    "total_layers":  0,
                    "lista_layers":  "",
                    "detalle_error": "",
                }

                layers, err = motor.leer_layers_con_timeout(dwg_path)
                motor.cerrar_docs_abiertos()

                if layers is None:
                    if not motor.vivo():
                        log.error("AutoCAD dejó de responder. Guardando progreso...")
                        checkpoint_guardar(ARCHIVO_CHECKPOINT, completados,
                                           vehiculo, filas)
                        return None
                    fila["detalle_error"] = err or "No se pudo leer"
                    log.warn(f"      ERROR: {nombre}  [{fila['detalle_error'][:50]}]")
                else:
                    ev = evaluar_estado(layers)
                    fila.update(ev)
                    fila["total_layers"] = len(layers)
                    fila["lista_layers"] = " | ".join(
                        f"{n}({i['color_texto']})"
                        for n, i in sorted(layers.items())
                    )
                    icono = {"ACTUALIZADA": "✓", "VIEJA": "✗", "INCOMPLETA": "!"}.get(ev["estado"], "?")
                    log.info(
                        f"      {icono} {ev['estado']:<14}  "
                        f"K={'azul✓' if ev['k_azul'] else ('no azul✗' if ev['tiene_k'] else 'NO')}  "
                        f"K2={'verde✓' if ev['k2_verde'] else ('no verde✗' if ev['tiene_k2'] else 'NO')}  "
                        f"{nombre}"
                    )

                filas.append(fila)
                rutas_ya_procesadas.add(dwg_path)
                nuevos_desde_guardado += 1

                # Guardar checkpoint cada N archivos
                if nuevos_desde_guardado >= GUARDAR_CADA_N:
                    checkpoint_guardar(ARCHIVO_CHECKPOINT, completados,
                                       vehiculo, filas)
                    log.info(f"      [checkpoint] {len(filas)} DWG(s) guardados en {vehiculo}")
                    nuevos_desde_guardado = 0

    return filas


# ──────────────────────────────────────────────────────────
# ESCANEO PRINCIPAL CON CHECKPOINT
# ──────────────────────────────────────────────────────────
def escanear(ruta_base, motor, completados, vehiculo_parcial, filas_parciales):
    """
    Recorre todos los vehículos con checkpoint por archivo y por vehículo.
    Retorna (datos_completos, completado_sin_errores)
    """
    if not os.path.exists(ruta_base):
        log.error(f"Ruta base no accesible: {ruta_base}")
        return completados, False

    vehiculos = sorted(d for d in os.listdir(ruta_base)
                       if os.path.isdir(os.path.join(ruta_base, d)))

    ya_listos  = [v for v in vehiculos if v in completados and v != vehiculo_parcial]
    pendientes = [v for v in vehiculos if v not in completados]
    # Si había uno en curso, asegurarse de que esté primero en pendientes
    if vehiculo_parcial and vehiculo_parcial in pendientes:
        pendientes = [vehiculo_parcial] + [v for v in pendientes if v != vehiculo_parcial]
    elif vehiculo_parcial and vehiculo_parcial not in completados:
        pendientes = [vehiculo_parcial] + pendientes

    log.info(f"Vehículos totales    : {len(vehiculos)}")
    log.info(f"Ya completados       : {len(ya_listos)}")
    if vehiculo_parcial:
        log.info(f"En curso (parcial)   : {vehiculo_parcial} "
                 f"({len(filas_parciales)} DWG(s) ya procesados)")
    log.info(f"Pendientes           : {len(pendientes)}")

    for idx, vehiculo in enumerate(pendientes, 1):
        ruta_vehiculo = os.path.join(ruta_base, vehiculo)
        log.info("=" * 70)
        log.info(f"[{idx}/{len(pendientes)}] VEHÍCULO: {vehiculo}")

        # Si es el vehículo parcial, retomar desde donde estaba
        fp = filas_parciales if vehiculo == vehiculo_parcial else []

        filas = procesar_vehiculo(vehiculo, ruta_vehiculo, motor,
                                  completados, filas_previas=fp)

        if filas is None:
            log.error("Proceso interrumpido por caída de AutoCAD.")
            log.error(f"Progreso guardado en: {ARCHIVO_CHECKPOINT}")
            log.error("Vuelve a abrir AutoCAD y ejecuta el script de nuevo para continuar.")
            return completados, False

        completados[vehiculo] = filas
        act  = sum(1 for f in filas if f["estado"] == "ACTUALIZADA")
        viej = sum(1 for f in filas if f["estado"] == "VIEJA")
        inc  = sum(1 for f in filas if f["estado"] == "INCOMPLETA")
        err  = sum(1 for f in filas if f["estado"] == "ERROR")
        log.ok(f"Vehículo completo: {vehiculo} — "
               f"{len(filas)} DWG(s)  ✓{act} ✗{viej} !{inc} E{err}")

        # Vehículo terminado: guardar como completo y limpiar en_curso
        checkpoint_guardar(ARCHIVO_CHECKPOINT, completados)

    return completados, True


# ──────────────────────────────────────────────────────────
# EXCEL — estilos
# ──────────────────────────────────────────────────────────
C = {
    "hdr":         "1F3864",
    "titulo":      "2E75B6",
    "white":       "FFFFFF",
    "alt":         "EEF4FF",
    "actualizada": "C6EFCE",
    "vieja":       "FFCCCC",
    "incompleta":  "FFEB9C",
    "error":       "FFC7CE",
}
_thin   = Side(style="thin", color="BBBBBB")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _ch(cell, bg=None):
    cell.font      = Font(name="Arial", bold=True, color=C["white"], size=10)
    cell.fill      = PatternFill("solid", start_color=bg or C["hdr"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border

def _cd(cell, alt=False, center=False, bold=False, bg=None, color_txt=None, mono=False):
    cell.font      = Font(name="Courier New" if mono else "Arial",
                          size=8 if mono else 9, bold=bold,
                          color=color_txt or "000000")
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border    = _border
    fill_bg = bg or (C["alt"] if alt else None)
    if fill_bg:
        cell.fill = PatternFill("solid", start_color=fill_bg)


ESTADO_CFG = {
    "ACTUALIZADA": {"bg": C["actualizada"], "label": "ACTUALIZADA"},
    "VIEJA":       {"bg": C["vieja"],       "label": "VIEJA / OBSOLETA"},
    "INCOMPLETA":  {"bg": C["incompleta"],  "label": "INCOMPLETA"},
    "ERROR":       {"bg": C["error"],       "label": "ERROR AL ABRIR"},
}

HEADERS = [
    "#", "VEHÍCULO", "MODELO", "VERSIÓN", "ARCHIVO", "ORIGEN",
    "ESTADO",
    "TIENE K", "NOMBRE K", "COLOR K", "K ES AZUL",
    "TIENE K2", "NOMBRE K2", "COLOR K2", "K2 ES VERDE",
    "TOTAL LAYERS", "LAYERS PRESENTES", "RUTA COMPLETA", "DETALLE ERROR",
]
ANCHOS = [5, 25, 25, 20, 35, 12, 18,
          10, 12, 14, 11,
          10, 12, 14, 12,
          14, 65, 80, 40]

def _titulo(ws, texto, ncols, fila=1, bg=None):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    c = ws.cell(fila, 1, texto)
    c.font      = Font(name="Arial", size=13, bold=True, color=C["white"])
    c.fill      = PatternFill("solid", start_color=bg or C["hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 30

def _meta(ws, texto, ncols, fila=2):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    c = ws.cell(fila, 1, texto)
    c.font      = Font(name="Arial", size=9, italic=True, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 16

def _headers(ws, fila=3, bg=None):
    for col, h in enumerate(HEADERS, 1):
        _ch(ws.cell(fila, col, h), bg=bg or C["titulo"])
    ws.row_dimensions[fila].height = 22

def _anchos(ws):
    for i, w in enumerate(ANCHOS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

def _fila_detalle(ws, row_excel, num, f):
    alt = (row_excel % 2 == 0)
    cfg = ESTADO_CFG.get(f["estado"], ESTADO_CFG["ERROR"])

    def yn(val):    return ("SÍ",   "C6EFCE") if val else ("NO",   "FFCCCC")
    def ytick(val): return ("SÍ ✓", "C6EFCE") if val else ("NO ✗", "FFCCCC")

    vals = [
        (num,                dict(center=True, alt=alt)),
        (f["vehiculo"],      dict(alt=alt)),
        (f["modelo"],        dict(alt=alt)),
        (f["version"],       dict(alt=alt)),
        (f["archivo"],       dict(alt=alt)),
        (f["origen"],        dict(center=True, alt=alt)),
        (cfg["label"],       dict(center=True, bold=True, bg=cfg["bg"])),
        yn(f["tiene_k"]),
        (f["nombre_k"],      dict(alt=alt)),
        (f["color_k"],       dict(alt=alt)),
        ytick(f["k_azul"]),
        yn(f["tiene_k2"]),
        (f["nombre_k2"],     dict(alt=alt)),
        (f["color_k2"],      dict(alt=alt)),
        ytick(f["k2_verde"]),
        (f["total_layers"],  dict(center=True, alt=alt)),
        (f["lista_layers"],  dict(alt=alt, mono=True)),
        (f["ruta"],          dict(alt=alt, mono=True, color_txt="0070C0")),
        (f["detalle_error"], dict(alt=alt, bg=C["error"] if f["estado"] == "ERROR" else None)),
    ]

    for col, item in enumerate(vals, 1):
        if isinstance(item, tuple) and len(item) == 2 and isinstance(item[1], str):
            _cd(ws.cell(row_excel, col, item[0]), center=True, bg=item[1])
        else:
            _cd(ws.cell(row_excel, col, item[0]), **item[1])


def crear_excel(datos, ruta_salida, parcial=False):
    log.info(f"Generando Excel{'  (PARCIAL — datos hasta ahora)' if parcial else ''}...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    todos  = [f for fs in datos.values() for f in fs]
    fecha  = datetime.now().strftime("%d/%m/%Y %H:%M")
    n      = len(HEADERS)
    t_act  = sum(1 for f in todos if f["estado"] == "ACTUALIZADA")
    t_viej = sum(1 for f in todos if f["estado"] == "VIEJA")
    t_inc  = sum(1 for f in todos if f["estado"] == "INCOMPLETA")
    t_err  = sum(1 for f in todos if f["estado"] == "ERROR")
    label_parcial = "  ⚠ PARCIAL" if parcial else ""

    # ── RESUMEN GENERAL ──────────────────────────────────
    ws_r = wb.create_sheet("RESUMEN GENERAL")
    ws_r.sheet_view.showGridLines = False
    _titulo(ws_r, f"AUDITORÍA LAYERS K / K2 — AGP PLANOS TÉCNICOS{label_parcial}", 8,
            bg="8B0000" if parcial else None)
    _meta(ws_r,
          f"Generado: {fecha}   |   Vehículos procesados: {len(datos)}   |   "
          f"Total DWG: {len(todos)}   |   "
          f"✓ {t_act}  ✗ {t_viej}  ! {t_inc}  E {t_err}", 8)

    RES_H = ["VEHÍCULO", "TOTAL DWG", "ACTUALIZADAS", "% ACT.",
             "VIEJAS", "% VIEJ.", "INCOMPLETAS", "ERRORES"]
    for col, h in enumerate(RES_H, 1):
        _ch(ws_r.cell(3, col, h))
    ws_r.row_dimensions[3].height = 22

    fila = 4
    for vehiculo, filas in sorted(datos.items()):
        act  = sum(1 for f in filas if f["estado"] == "ACTUALIZADA")
        viej = sum(1 for f in filas if f["estado"] == "VIEJA")
        inc  = sum(1 for f in filas if f["estado"] == "INCOMPLETA")
        err  = sum(1 for f in filas if f["estado"] == "ERROR")
        tot  = len(filas)
        alt  = fila % 2 == 0
        _cd(ws_r.cell(fila, 1, vehiculo), alt=alt, bold=True)
        _cd(ws_r.cell(fila, 2, tot),  alt=alt, center=True)
        _cd(ws_r.cell(fila, 3, act),  center=True, bg=C["actualizada"] if act  else (C["alt"] if alt else None))
        _cd(ws_r.cell(fila, 4, f"{act/tot*100:.1f}%" if tot else "0%"), alt=alt, center=True)
        _cd(ws_r.cell(fila, 5, viej), center=True, bg=C["vieja"]       if viej else (C["alt"] if alt else None))
        _cd(ws_r.cell(fila, 6, f"{viej/tot*100:.1f}%" if tot else "0%"), alt=alt, center=True)
        _cd(ws_r.cell(fila, 7, inc),  center=True, bg=C["incompleta"]  if inc  else (C["alt"] if alt else None))
        _cd(ws_r.cell(fila, 8, err),  center=True, bg=C["error"]       if err  else (C["alt"] if alt else None))
        fila += 1

    _cd(ws_r.cell(fila, 1, "TOTAL GENERAL"), bold=True)
    _cd(ws_r.cell(fila, 2, len(todos)), bold=True, center=True)
    _cd(ws_r.cell(fila, 3, t_act),  bold=True, center=True, bg=C["actualizada"])
    _cd(ws_r.cell(fila, 4, f"{t_act/len(todos)*100:.1f}%" if todos else "0%"), bold=True, center=True)
    _cd(ws_r.cell(fila, 5, t_viej), bold=True, center=True, bg=C["vieja"])
    _cd(ws_r.cell(fila, 6, f"{t_viej/len(todos)*100:.1f}%" if todos else "0%"), bold=True, center=True)
    _cd(ws_r.cell(fila, 7, t_inc),  bold=True, center=True, bg=C["incompleta"])
    _cd(ws_r.cell(fila, 8, t_err),  bold=True, center=True, bg=C["error"] if t_err else None)
    for i, w in enumerate([35, 12, 15, 10, 12, 10, 14, 10], 1):
        ws_r.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws_r.freeze_panes = "A4"

    # ── TODOS LOS ARCHIVOS ───────────────────────────────
    ws_t = wb.create_sheet("TODOS LOS ARCHIVOS")
    ws_t.sheet_view.showGridLines = False
    _titulo(ws_t, f"DETALLE COMPLETO — TODOS LOS VEHÍCULOS{label_parcial}", n,
            bg="8B0000" if parcial else None)
    _meta(ws_t, f"Generado: {fecha}   |   {len(todos)} archivos   |   Ruta base: {RUTA_BASE}", n)
    _headers(ws_t)
    for i, f in enumerate(todos, 1):
        _fila_detalle(ws_t, i + 3, i, f)
    _anchos(ws_t)
    ws_t.freeze_panes = "A4"

    # ── UNA HOJA POR VEHÍCULO ────────────────────────────
    for vehiculo, filas in sorted(datos.items()):
        nombre_hoja = (vehiculo[:31]
                       .replace("/","-").replace("\\","-")
                       .replace("*","").replace("?","")
                       .replace("[","").replace("]","").replace(":",""))
        ws_v = wb.create_sheet(nombre_hoja)
        ws_v.sheet_view.showGridLines = False
        act  = sum(1 for f in filas if f["estado"] == "ACTUALIZADA")
        viej = sum(1 for f in filas if f["estado"] == "VIEJA")
        inc  = sum(1 for f in filas if f["estado"] == "INCOMPLETA")
        err  = sum(1 for f in filas if f["estado"] == "ERROR")
        _titulo(ws_v, f"VEHÍCULO: {vehiculo}", n)
        _meta(ws_v,
              f"Total: {len(filas)}   |   Act.: {act}   Viejas: {viej}   "
              f"Incompletas: {inc}   Errores: {err}   |   {fecha}", n)
        _headers(ws_v)
        for i, f in enumerate(filas, 1):
            _fila_detalle(ws_v, i + 3, i, f)
        _anchos(ws_v)
        ws_v.freeze_panes = "A4"

    # ── REQUIEREN ATENCIÓN ───────────────────────────────
    problemas = [f for f in todos if f["estado"] != "ACTUALIZADA"]
    if problemas:
        ws_p = wb.create_sheet("REQUIEREN ATENCIÓN")
        ws_p.sheet_view.showGridLines = False
        _titulo(ws_p, f"ARTES QUE REQUIEREN ATENCIÓN — {len(problemas)} archivos", n, bg="C00000")
        _meta(ws_p,
              f"VIEJAS: {t_viej}   INCOMPLETAS: {t_inc}   ERRORES: {t_err}   |   {fecha}", n)
        _headers(ws_p, bg="C00000")
        for i, f in enumerate(problemas, 1):
            _fila_detalle(ws_p, i + 3, i, f)
        _anchos(ws_p)
        ws_p.freeze_panes = "A4"

    wb.save(ruta_salida)
    log.ok(f"Excel guardado: {ruta_salida}")


# ──────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Auditoría Layers K/K2")
    parser.add_argument("--reiniciar",   action="store_true",
                        help="Borra el checkpoint y empieza desde cero")
    parser.add_argument("--solo-excel",  action="store_true",
                        help="No escanea, solo genera el Excel desde el checkpoint existente")
    args = parser.parse_args()

    log.info("=" * 70)
    log.info("  AUDITORÍA LAYERS K / K2 — AGP PLANOS TÉCNICOS")
    log.info("=" * 70)
    log.info(f"Ruta base  : {RUTA_BASE}")
    log.info(f"Checkpoint : {ARCHIVO_CHECKPOINT}")
    log.info("-" * 70)

    # ── Modo: solo generar Excel desde checkpoint ──────
    if args.solo_excel:
        completados, _, _ = checkpoint_cargar(ARCHIVO_CHECKPOINT)
        if not completados:
            log.error("No hay checkpoint guardado. Ejecuta sin --solo-excel primero.")
            return
        nombre = f"Auditoria_Layers_K_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        crear_excel(completados, nombre)
        return

    # ── Reiniciar desde cero ───────────────────────────
    if args.reiniciar:
        checkpoint_borrar(ARCHIVO_CHECKPOINT)
        log.info("Checkpoint borrado. Empezando desde cero.")

    # ── Cargar progreso previo ─────────────────────────
    completados, vehiculo_parcial, filas_parciales = checkpoint_cargar(ARCHIVO_CHECKPOINT)
    hay_progreso = bool(completados or vehiculo_parcial)

    if hay_progreso:
        resp = input("  ¿Continuar desde donde se quedó? [S/n]: ").strip().lower()
        if resp == "n":
            checkpoint_borrar(ARCHIVO_CHECKPOINT)
            completados, vehiculo_parcial, filas_parciales = {}, None, []
            log.info("Checkpoint descartado. Empezando desde cero.")

    log.info("\nIMPORTANTE: AutoCAD debe estar abierto (sin archivos) antes de continuar.")
    input("  Presiona Enter cuando AutoCAD esté listo...")

    motor = AutoCAD()

    t0 = time.time()
    datos, completado = escanear(RUTA_BASE, motor,
                                 completados, vehiculo_parcial, filas_parciales)
    motor.quit()
    duracion = time.time() - t0

    if not datos:
        log.warn("No se encontraron datos.")
        return

    todos  = [f for fs in datos.values() for f in fs]
    t_act  = sum(1 for f in todos if f["estado"] == "ACTUALIZADA")
    t_viej = sum(1 for f in todos if f["estado"] == "VIEJA")
    t_inc  = sum(1 for f in todos if f["estado"] == "INCOMPLETA")
    t_err  = sum(1 for f in todos if f["estado"] == "ERROR")

    nombre_excel = f"Auditoria_Layers_K_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    crear_excel(datos, nombre_excel, parcial=not completado)

    if completado:
        log.ok("Proceso completado. Borrando checkpoint...")
        checkpoint_borrar(ARCHIVO_CHECKPOINT)
    else:
        log.warn(f"Proceso interrumpido. Checkpoint guardado: {ARCHIVO_CHECKPOINT}")
        log.warn("Vuelve a ejecutar para continuar desde donde se quedó.")

    log.info("\n" + "=" * 70)
    log.info("  RESUMEN FINAL")
    log.info("=" * 70)
    log.info(f"  Vehículos procesados  : {len(datos)}")
    log.info(f"  Total DWG             : {len(todos)}")
    log.info(f"  ACTUALIZADAS          : {t_act}")
    log.info(f"  VIEJAS                : {t_viej}")
    log.info(f"  INCOMPLETAS           : {t_inc}")
    log.info(f"  ERRORES               : {t_err}")
    h = int(duracion//3600); m = int((duracion%3600)//60); s = int(duracion%60)
    log.info(f"  Tiempo esta sesión    : {h}h {m}m {s}s")
    log.info(f"  Excel                 : {nombre_excel}")
    if not completado:
        log.warn("  ⚠ EXCEL PARCIAL — falta procesar vehículos pendientes")
    log.info("=" * 70)


if __name__ == "__main__":
    main()
