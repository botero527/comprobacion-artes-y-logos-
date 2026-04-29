#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
AUDITORÍA DE LOGOS — ALFA ROMEO
Procesa: \\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS\ALFA ROMEO
Estructura: MARCA > VEHICULO > VERSION (V*) > ARTES > DWGs
"""

import os
import sys
import time
from collections import defaultdict

class Logger:
    def _fmt(self, msg):
        return f"{time.strftime('%H:%M:%S')}  {msg}"
    def info(self, msg):    print(self._fmt(msg))
    def warn(self, msg):    print(self._fmt(f"[WARN]  {msg}"))
    def error(self, msg):   print(self._fmt(f"[ERROR] {msg}"))

log = Logger()

try:
    import win32com.client
    import pythoncom
except ImportError:
    print("Falta pywin32. Ejecuta: pip install pywin32")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)


RUTA_BASE     = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS\ALFA ROMEO"
ARCHIVO_EXCEL = "Auditoria_ALFA_ROMEO.xlsx"
PATRONES_LOGO = ["logo", "trazabilidad"]

C = {
    "hdr":    "1F4E78",
    "titulo": "2E75B6",
    "ok":     "C6EFCE",
    "falta":  "FFCCCC",
    "error":  "FFC7CE",
    "white":  "FFFFFF",
}

ESTADOS = {
    "OK":    {"color": C["ok"],    "texto": "Tiene Logo"},
    "FALTA": {"color": C["falta"], "texto": "Falta Logo"},
    "ERROR": {"color": C["error"], "texto": "Error"},
}


class AutoCADMotor:
    def __init__(self):
        self.acad = None
        self._conectar()

    def _conectar(self):
        pythoncom.CoInitialize()
        try:
            self.acad = win32com.client.GetActiveObject("AutoCAD.Application")
            log.info("[AutoCAD] Conectado a instancia existente")
        except Exception:
            log.error("[ERROR] No hay AutoCAD abierto. Abre AutoCAD primero y vuelve a ejecutar.")
            sys.exit(1)

    def abrir(self, ruta):
        ruta_abs = os.path.abspath(ruta)
        tiempos = [1.0, 2.0, 3.0, 4.0, 5.0]
        
        for i, espera in enumerate(tiempos):
            try:
                doc = self.acad.Documents.Open(ruta_abs)
                time.sleep(1.0)
                return doc
            except Exception as e:
                if i < len(tiempos) - 1:
                    log.warn(f"  Reintento {i+1}/{len(tiempos)} — {str(e)[:30]}")
                    time.sleep(espera)
                else:
                    return None
        return None

    def cerrar(self, doc):
        try:
            doc.Close(False)
        except Exception:
            pass

    def tiene_layer_logo(self, doc):
        try:
            layers = doc.Layers
            for i in range(layers.Count):
                try:
                    nombre = layers.Item(i).Name
                    nombre_upper = nombre.upper()
                    for patron in PATRONES_LOGO:
                        if patron.upper() in nombre_upper:
                            return True, nombre
                except Exception:
                    continue
            return False, ""
        except Exception as e:
            return None, str(e)

    def quit(self):
        # No cierra AutoCAD, solo se desconecta
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def recolectar_dwgs_de_artes(ruta_artes):
    dwgs = []
    try:
        items = os.listdir(ruta_artes)
    except Exception:
        return dwgs

    for item in items:
        ruta_item = os.path.join(ruta_artes, item)

        if os.path.isfile(ruta_item) and item.lower().endswith(".dwg"):
            dwgs.append(ruta_item)

        elif os.path.isdir(ruta_item) and item.upper() == "BN":
            try:
                for archivo in os.listdir(ruta_item):
                    ruta_arch = os.path.join(ruta_item, archivo)
                    if os.path.isfile(ruta_arch) and archivo.lower().endswith(".dwg"):
                        dwgs.append(ruta_arch)
            except Exception:
                pass

    return sorted(set(dwgs))


def analizar_arte(motor, arte_path, vehiculo, version):
    nombre = os.path.basename(arte_path)
    
    estado  = "ERROR"
    detalle = "No se pudo abrir"
    layer   = ""

    doc = motor.abrir(arte_path)
    if doc is None:
        detalle = "No se pudo abrir"
        log.warn(f"       ✗ {nombre}")
    else:
        try:
            tiene, layer_nombre = motor.tiene_layer_logo(doc)
            
            if tiene is True:
                estado  = "OK"
                layer   = layer_nombre
                detalle = f"Layer: {layer_nombre}"
                log.info(f"       ✓ {layer_nombre}")
            elif tiene is False:
                estado  = "FALTA"
                layer   = ""
                detalle = "Sin layer de logo"
                log.info(f"       ✗ Sin logo")
            else:
                estado  = "ERROR"
                layer   = ""
                detalle = layer_nombre
                log.error(f"       ERROR: {detalle}")
        except Exception as e:
            estado  = "ERROR"
            detalle = str(e)[:50]
            log.error(f"       ERROR: {detalle}")
        finally:
            motor.cerrar(doc)
    
    time.sleep(0.3)
    
    return {
        "vehiculo": vehiculo,
        "version":  version,
        "arte":     nombre,
        "ruta":     arte_path,
        "estado":   estado,
        "detalle":  detalle,
        "layer":    layer,
    }


def escanear(ruta_base, motor):
    datos = defaultdict(list)

    if not os.path.exists(ruta_base):
        log.error(f"Ruta base no existe: {ruta_base}")
        return datos

    vehiculos = sorted(
        d for d in os.listdir(ruta_base)
        if os.path.isdir(os.path.join(ruta_base, d))
    )
    
    log.info(f"Vehículos encontrados: {len(vehiculos)}")

    for vehiculo in vehiculos:
        ruta_vehiculo = os.path.join(ruta_base, vehiculo)

        versiones = sorted(
            d for d in os.listdir(ruta_vehiculo)
            if os.path.isdir(os.path.join(ruta_vehiculo, d))
            and d.upper().startswith("V")
        )

        if not versiones:
            log.info(f"  {vehiculo}: sin versiones")
            continue

        log.info("=" * 60)
        log.info(f"VEHICULO: {vehiculo} ({len(versiones)} versiones)")
        log.info("=" * 60)

        for version in versiones:
            ruta_version = os.path.join(ruta_vehiculo, version)

            ruta_artes = None
            try:
                for sub in os.listdir(ruta_version):
                    if sub.upper() == "ARTES" and \
                       os.path.isdir(os.path.join(ruta_version, sub)):
                        ruta_artes = os.path.join(ruta_version, sub)
                        break
            except Exception:
                continue

            if not ruta_artes:
                log.info(f"  {version}: sin carpeta ARTES")
                continue

            dwgs = recolectar_dwgs_de_artes(ruta_artes)

            if not dwgs:
                log.info(f"    Sin DWGs en ARTES")
                continue

            log.info(f"    {version}: {len(dwgs)} archivos")

            for dwg_path in dwgs:
                log.info(f"    -> {os.path.basename(dwg_path)}")
                fila = analizar_arte(motor, dwg_path, vehiculo, version)
                datos[vehiculo].append(fila)

    return datos


def _header_row(ws, headers, fill_color=None):
    fill_color = fill_color or C["titulo"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color=C["white"])
        cell.fill = PatternFill("solid", start_color=fill_color)
        cell.alignment = Alignment(horizontal="center")


def crear_excel(datos, ruta_salida):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    todos = []
    for vehiculo, filas in datos.items():
        todos.extend(filas)

    if not todos:
        log.warn("No hay datos para generar Excel")
        return

    # RESUMEN
    ws = wb.create_sheet("RESUMEN")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "AUDITORÍA DE LOGOS — ALFA ROMEO"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=C["white"])
    ws["A1"].fill = PatternFill("solid", start_color=C["hdr"])
    ws["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Vehículo", "Con Logo", "Sin Logo", "Errores", "Total"], 1):
        cell = ws.cell(3, col, h)
        cell.font = Font(bold=True, color=C["white"])
        cell.fill = PatternFill("solid", start_color=C["titulo"])

    row = 4
    for vehiculo, filas in sorted(datos.items()):
        ok    = sum(1 for f in filas if f["estado"] == "OK")
        falta = sum(1 for f in filas if f["estado"] == "FALTA")
        error = sum(1 for f in filas if f["estado"] == "ERROR")
        
        ws.cell(row, 1, vehiculo)
        ws.cell(row, 2, ok)
        ws.cell(row, 3, falta)
        ws.cell(row, 4, error)
        ws.cell(row, 5, ok + falta + error)
        row += 1

    ok_t    = sum(1 for f in todos if f["estado"] == "OK")
    falta_t = sum(1 for f in todos if f["estado"] == "FALTA")
    error_t = sum(1 for f in todos if f["estado"] == "ERROR")
    for col, val in enumerate(["TOTAL", ok_t, falta_t, error_t, len(todos)], 1):
        ws.cell(row, col, val).font = Font(bold=True)

    ws.column_dimensions["A"].width = 40
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 12

    # TODOS
    ws = wb.create_sheet("TODOS")
    headers = ["Vehículo", "Versión", "Archivo", "Estado", "Detalle", "Layer", "Ruta"]
    _header_row(ws, headers)

    for row, f in enumerate(todos, 2):
        ws.cell(row, 1, f["vehiculo"])
        ws.cell(row, 2, f["version"])
        ws.cell(row, 3, f["arte"])
        estado_cell = ws.cell(row, 4, ESTADOS[f["estado"]]["texto"])
        estado_cell.fill = PatternFill("solid", start_color=ESTADOS[f["estado"]]["color"])
        ws.cell(row, 5, f["detalle"])
        ws.cell(row, 6, f["layer"])
        ws.cell(row, 7, f["ruta"])

    anchos = [40, 30, 40, 15, 35, 25, 70]
    for col, w in zip(["A","B","C","D","E","F","G"], anchos):
        ws.column_dimensions[col].width = w

    wb.save(ruta_salida)
    log.info(f"Excel guardado: {ruta_salida}")


def main():
    log.info("=" * 60)
    log.info("AUDITORÍA DE LOGOS — ALFA ROMEO")
    log.info("=" * 60)
    log.info(f"Ruta base : {RUTA_BASE}")
    log.info(f"Patrones  : {PATRONES_LOGO}")
    log.info("-" * 60)

    log.info("\n[1/3] Iniciando AutoCAD...")
    motor = AutoCADMotor()

    log.info("\n[2/3] Escaneando archivos DWG...")
    inicio = time.time()
    datos = escanear(RUTA_BASE, motor)
    duracion = time.time() - inicio

    if not datos:
        log.warn("No se encontraron archivos.")
        motor.quit()
        return

    log.info("\n[3/3] Generando Excel...")
    crear_excel(datos, ARCHIVO_EXCEL)

    todos = []
    for filas in datos.values():
        todos.extend(filas)

    ok      = sum(1 for f in todos if f["estado"] == "OK")
    falta   = sum(1 for f in todos if f["estado"] == "FALTA")
    error   = sum(1 for f in todos if f["estado"] == "ERROR")

    log.info("\n" + "=" * 60)
    log.info("COMPLETADO")
    log.info(f"  Con logo  : {ok}")
    log.info(f"  Sin logo  : {falta}")
    log.info(f"  Errores   : {error}")
    log.info(f"  Total     : {len(todos)}")
    log.info(f"  Tiempo    : {duracion:.1f} segundos")
    log.info(f"  Excel     : {ARCHIVO_EXCEL}")
    log.info("=" * 60)

    motor.quit()


if __name__ == "__main__":
    main()