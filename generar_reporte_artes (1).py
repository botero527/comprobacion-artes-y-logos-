"""
Script: Generador de Reporte de Rutas ARTES - AGP (Para migración IT)
Uso: python generar_reporte_artes.py
Requiere: pip install openpyxl
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================
# CONFIGURACION
# ============================================================
RUTA_BASE = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS\CHEVROLET"
NOMBRE_CARPETA_ARTE = "ARTES"
ARCHIVO_SALIDA = "Migracion_ARTES_AGP_2.xlsx"
# ============================================================

COLOR_HEADER   = "1F3864"
COLOR_TITULO   = "2E75B6"
COLOR_BLANCO   = "FFFFFF"
COLOR_FILA_ALT = "EEF4FF"

thin   = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def estilo_header(cell, bg=COLOR_HEADER):
    cell.font      = Font(name="Arial", bold=True, color=COLOR_BLANCO, size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border


def estilo_dato(cell, alt=False, center=False):
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border    = border
    if alt:
        cell.fill = PatternFill("solid", start_color=COLOR_FILA_ALT)


def obtener_peso_carpeta(ruta):
    """
    Calcula el tamaño total de una carpeta (incluyendo subcarpetas).
    Retorna el tamaño en bytes.
    """
    total_size = 0
    try:
        for dirpath, dirnames, filenames in os.walk(ruta):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                try:
                    total_size += os.path.getsize(filepath)
                except (OSError, FileNotFoundError):
                    # Si no se puede obtener el tamaño, continuar para evitar errores 
                    pass
    except (OSError, FileNotFoundError):
        pass
    return total_size


def formato_peso(bytes_peso):
    """
    Convierte bytes a formato legible (KB, MB, GB).
    """
    for unidad in ['B', 'KB', 'MB', 'GB', 'TB']:
        if bytes_peso < 1024.0:
            return f"{bytes_peso:.2f} {unidad}"
        bytes_peso /= 1024.0
    return f"{bytes_peso:.2f} PB"


def escanear(ruta_base):
    """
    Recorre: Vehiculo > Modelo > Version
    En cada Version busca carpeta ARTES y guarda la ruta exacta UNC.
    Retorna dict: { vehiculo: [ {modelo, version, ruta_arte, peso} ] }
    """
    resultado = {}

    if not os.path.exists(ruta_base):
        print(f"[ERROR] No se puede acceder a: {ruta_base}")
        print("Verifica conexion de red y que la ruta este accesible.")
        return resultado

    vehiculos = sorted([d for d in os.listdir(ruta_base)
                        if os.path.isdir(os.path.join(ruta_base, d))])

    print(f"Vehiculos encontrados: {len(vehiculos)}\n")

    for vehiculo in vehiculos:
        ruta_vehiculo = os.path.join(ruta_base, vehiculo)
        filas = []

        modelos = sorted([d for d in os.listdir(ruta_vehiculo)
                          if os.path.isdir(os.path.join(ruta_vehiculo, d))])

        for modelo in modelos:
            ruta_modelo = os.path.join(ruta_vehiculo, modelo)

            versiones = sorted([d for d in os.listdir(ruta_modelo)
                                 if os.path.isdir(os.path.join(ruta_modelo, d))])

            for version in versiones:
                ruta_version = os.path.join(ruta_vehiculo, modelo, version)
                ruta_arte    = os.path.join(ruta_version, NOMBRE_CARPETA_ARTE)

                if os.path.isdir(ruta_arte):
                    # Calcular peso de la carpeta ARTES
                    peso_bytes = obtener_peso_carpeta(ruta_arte)
                    peso_formateado = formato_peso(peso_bytes)
                    
                    filas.append({
                        "modelo":      modelo,
                        "version":     version,
                        "ruta_arte":   ruta_arte,
                        "peso_bytes":  peso_bytes,
                        "peso":        peso_formateado,
                    })
                    print(f"  [OK] {vehiculo} | {modelo} | {version} - {peso_formateado}")
                else:
                    print(f"  [--] {vehiculo} | {modelo} | {version} (sin ARTES)")

        if filas:
            resultado[vehiculo] = filas

    return resultado


def crear_excel(resultado):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_rutas = sum(len(v) for v in resultado.values())

    # Calcular peso total de todas las carpetas (para mostrar en encabezado)
    peso_total_bytes = 0
    for filas in resultado.values():
        for datos in filas:
            peso_total_bytes += datos.get("peso_bytes", 0)

    # ===========================================================
    # HOJA MAESTRA — todas las rutas juntas para IT
    # ===========================================================
    ws_master = wb.create_sheet("TODAS LAS RUTAS (IT)")

    ws_master.merge_cells("A1:F1")
    ws_master["A1"] = "MIGRACION ARTES — AGP PLANOS TECNICOS"
    ws_master["A1"].font      = Font(name="Arial", size=14, bold=True, color=COLOR_BLANCO)
    ws_master["A1"].fill      = PatternFill("solid", start_color=COLOR_HEADER)
    ws_master["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_master.row_dimensions[1].height = 32

    ws_master.merge_cells("A2:F2")
    ws_master["A2"] = (f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
                       f"Total carpetas ARTES: {total_rutas}   |   "
                       f"Peso Total: {formato_peso(peso_total_bytes)}   |   "
                       f"Origen: {RUTA_BASE}")
    ws_master["A2"].font      = Font(name="Arial", size=9, italic=True, color="555555")
    ws_master["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws_master.row_dimensions[2].height = 18

    encabezados = ["#", "VEHICULO", "MODELO", "VERSION", "PESO (MB)", "RUTA CARPETA ARTES"]
    for col, h in enumerate(encabezados, 1):
        estilo_header(ws_master.cell(3, col, h))
    ws_master.row_dimensions[3].height = 20

    fila = 4
    for vehiculo, filas in sorted(resultado.items()):
        for datos in filas:
            alt = (fila % 2 == 0)
            peso_bytes = datos.get("peso_bytes", 0)
            peso_total_bytes += peso_bytes
            peso_mb = peso_bytes / (1024 * 1024)  # Convertir a MB
            
            estilo_dato(ws_master.cell(fila, 1, fila - 3), alt=alt, center=True)
            estilo_dato(ws_master.cell(fila, 2, vehiculo),          alt=alt)
            estilo_dato(ws_master.cell(fila, 3, datos["modelo"]),   alt=alt)
            estilo_dato(ws_master.cell(fila, 4, datos["version"]),  alt=alt)
            estilo_dato(ws_master.cell(fila, 5, f"{peso_mb:.2f}"),   alt=alt, center=True)

            c_ruta = ws_master.cell(fila, 6, datos["ruta_arte"])
            c_ruta.font      = Font(name="Courier New", size=9, color="0070C0")
            c_ruta.alignment = Alignment(horizontal="left", vertical="center")
            c_ruta.border    = border
            if alt:
                c_ruta.fill = PatternFill("solid", start_color=COLOR_FILA_ALT)

            fila += 1

    ws_master.column_dimensions["A"].width = 6
    ws_master.column_dimensions["B"].width = 25
    ws_master.column_dimensions["C"].width = 25
    ws_master.column_dimensions["D"].width = 25
    ws_master.column_dimensions["E"].width = 15
    ws_master.column_dimensions["F"].width = 80
    ws_master.freeze_panes = "A4"

    # ===========================================================
    # UNA HOJA POR VEHICULO
    # ===========================================================
    for vehiculo, filas in sorted(resultado.items()):
        nombre_hoja = (vehiculo[:31]
                       .replace("/","-").replace("\\","-")
                       .replace("*","").replace("?","")
                       .replace("[","").replace("]","").replace(":",""))
        ws = wb.create_sheet(nombre_hoja)

        # Calcular peso total del vehículo
        peso_vehiculo_bytes = sum(f.get("peso_bytes", 0) for f in filas)
        
        ws.merge_cells("A1:E1")
        ws["A1"] = f"VEHICULO: {vehiculo}"
        ws["A1"].font      = Font(name="Arial", size=13, bold=True, color=COLOR_BLANCO)
        ws["A1"].fill      = PatternFill("solid", start_color=COLOR_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:E2")
        ws["A2"] = (f"Total carpetas ARTES: {len(filas)}   |   "
                    f"Peso Total: {formato_peso(peso_vehiculo_bytes)}   |   "
                    f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="555555")
        ws["A2"].alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 16

        enc = ["MODELO", "VERSION", "PESO (MB)", "RUTA CARPETA ARTES"]
        for col, h in enumerate(enc, 1):
            estilo_header(ws.cell(3, col, h), bg=COLOR_TITULO)
        ws.row_dimensions[3].height = 20

        for i, datos in enumerate(filas, 1):
            row = i + 3
            alt = (i % 2 == 0)
            peso_bytes = datos.get("peso_bytes", 0)
            peso_mb = peso_bytes / (1024 * 1024)

            estilo_dato(ws.cell(row, 1, datos["modelo"]),  alt=alt)
            estilo_dato(ws.cell(row, 2, datos["version"]), alt=alt)
            estilo_dato(ws.cell(row, 3, f"{peso_mb:.2f}"), alt=alt, center=True)

            c_ruta = ws.cell(row, 4, datos["ruta_arte"])
            c_ruta.font      = Font(name="Courier New", size=9, color="0070C0")
            c_ruta.alignment = Alignment(horizontal="left", vertical="center")
            c_ruta.border    = border
            if alt:
                c_ruta.fill = PatternFill("solid", start_color=COLOR_FILA_ALT)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 80
        ws.freeze_panes = "A4"

    wb.save(ARCHIVO_SALIDA)
    print(f"\n{'='*60}")
    print(f"[OK] Excel generado: {ARCHIVO_SALIDA}")
    print(f"     Total carpetas ARTES mapeadas: {total_rutas}")
    print(f"     Vehiculos con ARTES: {len(resultado)}")
    print(f"     Peso Total: {formato_peso(peso_total_bytes)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    print("=" * 60)
    print("  REPORTE RUTAS ARTES — AGP PLANOS TECNICOS")
    print("=" * 60)
    print(f"Ruta base: {RUTA_BASE}\n")

    resultado = escanear(RUTA_BASE)

    if not resultado:
        print("\n[!] No se encontraron carpetas ARTES.")
        print("    Verifica conexion de red y ruta.")
    else:
        crear_excel(resultado)
