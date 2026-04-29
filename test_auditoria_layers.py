#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
TEST — Auditoría Layers K / K2
Versión de prueba: selecciona archivos DWG manualmente y valida la lógica
antes de correr el script completo sobre toda la red.
"""

import os
import sys
import time
from datetime import datetime

# ──────────────────────────────────────────────────────────
# ARCHIVOS A PROBAR — agrega/quita rutas aquí
# ──────────────────────────────────────────────────────────
ARCHIVOS_PRUEBA = [
    # r"\\192.168.2.37\ingenieria\...\archivo1.dwg",
    # r"\\192.168.2.37\ingenieria\...\archivo2.dwg",
    # r"C:\Users\abotero\Desktop\prueba.dwg",
]
# ──────────────────────────────────────────────────────────

COLOR_VERDE_ACI = 3
COLOR_AZUL_ACI  = 5

try:
    import win32com.client
    import pythoncom
except ImportError:
    print("Falta pywin32. Ejecuta: pip install pywin32")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────
# HELPERS VISUALES EN CONSOLA
# ──────────────────────────────────────────────────────────
def sep(char="─", n=65): print(char * n)

def log(msg, prefijo=""):
    ts = time.strftime("%H:%M:%S")
    print(f"{ts}  {prefijo}{msg}")

def ok(msg):    log(msg, "  ✓  ")
def warn(msg):  log(msg, "  !  ")
def err(msg):   log(msg, "  ✗  ")
def info(msg):  log(msg, "     ")


# ──────────────────────────────────────────────────────────
# AUTOCAD
# ──────────────────────────────────────────────────────────
def conectar_autocad():
    pythoncom.CoInitialize()
    try:
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
        ok("AutoCAD conectado")
        return acad
    except Exception:
        err("No hay AutoCAD abierto. Ábrelo primero.")
        sys.exit(1)


def abrir_dwg(acad, ruta):
    ruta_abs = os.path.abspath(ruta)
    for intento in range(1, 4):
        try:
            doc = acad.Documents.Open(ruta_abs)
            time.sleep(1.0)
            return doc
        except Exception as e:
            warn(f"  Reintento {intento}/3: {str(e)[:50]}")
            time.sleep(intento * 2.0)
    return None


def cerrar_dwg(doc):
    try:
        doc.Close(False)
    except Exception:
        pass


def leer_layers(doc):
    """Retorna lista de dicts con info de cada layer."""
    layers = []
    try:
        for i in range(doc.Layers.Count):
            try:
                l = doc.Layers.Item(i)
                layers.append({
                    "nombre":      l.Name,
                    "color_aci":   l.Color,
                    "color_texto": _nombre_color(l.Color),
                })
            except Exception:
                continue
    except Exception as e:
        warn(f"Error leyendo layers: {e}")
    return layers


def _nombre_color(aci):
    mapa = {
        1: "Rojo", 2: "Amarillo", 3: "Verde", 4: "Cyan",
        5: "Azul", 6: "Magenta",  7: "Blanco/Negro", 8: "Gris oscuro",
        9: "Gris claro", 0: "ByBlock", 256: "ByLayer",
    }
    return mapa.get(aci, f"Color ACI {aci}")


# ──────────────────────────────────────────────────────────
# LÓGICA DE VALIDACIÓN
# ──────────────────────────────────────────────────────────
def evaluar(layers):
    """
    Busca layers K y K2.
    Condiciones:
      - Layer K debe existir y ser AZUL (ACI 5)
      - Layer K2 debe existir y ser VERDE (ACI 3)
    """
    layer_k  = next((l for l in layers if l["nombre"].upper().strip() == "K"),  None)
    layer_k2 = next((l for l in layers if l["nombre"].upper().strip() == "K2"), None)

    tiene_k      = layer_k  is not None
    tiene_k2     = layer_k2 is not None
    k_es_azul    = tiene_k  and (layer_k["color_aci"]  == COLOR_AZUL_ACI)
    k2_es_verde  = tiene_k2 and (layer_k2["color_aci"] == COLOR_VERDE_ACI)

    if k_es_azul and k2_es_verde:
        estado = "ACTUALIZADA"
    elif not tiene_k and not tiene_k2:
        estado = "VIEJA"
    else:
        estado = "INCOMPLETA"

    return {
        "estado":      estado,
        "tiene_k":     tiene_k,
        "layer_k":     layer_k,
        "k_es_azul":   k_es_azul,
        "tiene_k2":    tiene_k2,
        "layer_k2":    layer_k2,
        "k2_es_verde": k2_es_verde,
    }


# ──────────────────────────────────────────────────────────
# ANÁLISIS DE UN ARCHIVO + CONSOLA DETALLADA
# ──────────────────────────────────────────────────────────
def analizar_archivo(acad, ruta, numero, total):
    nombre = os.path.basename(ruta)
    sep()
    info(f"[{numero}/{total}] {nombre}")
    info(f"Ruta: {ruta}")
    sep("·")

    if not os.path.exists(ruta):
        err("Archivo NO encontrado en disco")
        return {"archivo": nombre, "ruta": ruta, "estado": "ERROR",
                "detalle": "Archivo no existe", "layers": []}

    info("Abriendo en AutoCAD...")
    doc = abrir_dwg(acad, ruta)

    if doc is None:
        err("No se pudo abrir el archivo")
        return {"archivo": nombre, "ruta": ruta, "estado": "ERROR",
                "detalle": "No se pudo abrir", "layers": []}

    layers = leer_layers(doc)
    cerrar_dwg(doc)

    # Mostrar todos los layers en consola
    info(f"Total layers en el archivo: {len(layers)}")
    print()
    print("  {:<30} {:>10}  {}".format("NOMBRE LAYER", "COLOR ACI", "COLOR"))
    print("  " + "─" * 55)
    for l in sorted(layers, key=lambda x: x["nombre"]):
        marcador = ""
        n = l["nombre"].upper().strip()
        if n == "K":
            color_aviso = " (AZUL ✓)" if l["color_aci"] == COLOR_AZUL_ACI else f" (NO ES AZUL ✗ — es {l['color_texto']})"
            marcador = f"  ◄── LAYER K{color_aviso}"
        elif n == "K2":
            color_aviso = " (VERDE ✓)" if l["color_aci"] == COLOR_VERDE_ACI else f" (NO ES VERDE ✗ — es {l['color_texto']})"
            marcador = f"  ◄── LAYER K2{color_aviso}"
        print(f"  {l['nombre']:<30} {l['color_aci']:>10}  {l['color_texto']}{marcador}")

    print()
    diag = evaluar(layers)

    # Diagnóstico claro
    sep("═")
    ICONOS = {"ACTUALIZADA": "✓", "VIEJA": "✗", "INCOMPLETA": "!"}
    icono = ICONOS.get(diag["estado"], "?")
    print(f"\n  {icono}  RESULTADO: {diag['estado']}\n")

    print(f"  Layer K   : {'SÍ existe' if diag['tiene_k'] else 'NO existe'}", end="")
    if diag["layer_k"]:
        print(f"  → nombre='{diag['layer_k']['nombre']}' color={diag['layer_k']['color_texto']}", end="")
    print()
    if diag["tiene_k"]:
        print(f"  K es azul : {'SÍ ✓' if diag['k_es_azul'] else 'NO ✗  ← necesita ser AZUL (ACI 5)'}")

    print(f"  Layer K2  : {'SÍ existe' if diag['tiene_k2'] else 'NO existe'}", end="")
    if diag["layer_k2"]:
        print(f"  → nombre='{diag['layer_k2']['nombre']}' color={diag['layer_k2']['color_texto']}", end="")
    print()
    if diag["tiene_k2"]:
        print(f"  K2 verde  : {'SÍ ✓' if diag['k2_es_verde'] else 'NO ✗  ← necesita ser VERDE (ACI 3)'}")

    EXPLICACIONES = {
        "ACTUALIZADA":
            "  Tiene layer K  +  layer K2 en verde → ARTE ACTUALIZADA",
        "VIEJA":
            "  Sin layer K y sin K2 verde → ARTE VIEJA / OBSOLETA",
        "INCOMPLETA":
            "  Solo cumple una condición → revisar manualmente",
    }
    print(f"\n  {EXPLICACIONES[diag['estado']]}")
    print()

    return {
        "archivo": nombre,
        "ruta":    ruta,
        "estado":        diag["estado"],
        "detalle":       diag,
        "layers":        layers,
        "detalle_error": "",
    }


# ──────────────────────────────────────────────────────────
# REPORTE EXCEL DE PRUEBA
# ──────────────────────────────────────────────────────────
_thin   = Side(style="thin", color="BBBBBB")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

BG = {
    "ACTUALIZADA": "C6EFCE",
    "VIEJA":       "FFCCCC",
    "INCOMPLETA":  "FFEB9C",
    "ERROR":       "FFC7CE",
}

def _c(ws, r, col, val, bold=False, bg=None, center=False, color_txt="000000",
       fuente="Arial", size=10):
    cell = ws.cell(r, col, val)
    cell.font      = Font(name=fuente, bold=bold, size=size, color=color_txt)
    cell.fill      = PatternFill("solid", start_color=bg) if bg else PatternFill()
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=False)
    cell.border    = _border
    return cell


def generar_excel_prueba(resultados, ruta_salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RESULTADOS PRUEBA"
    ws.sheet_view.showGridLines = False

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Título
    ws.merge_cells("A1:Q1")
    t = ws.cell(1, 1, f"TEST AUDITORÍA LAYERS K/K2 — {fecha}")
    t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    HEADERS = [
        "#", "ARCHIVO", "ESTADO",
        "TIENE K", "NOMBRE K", "COLOR K", "K ES AZUL",
        "TIENE K2", "NOMBRE K2", "COLOR K2", "K2 ES VERDE",
        "TOTAL LAYERS", "TODOS LOS LAYERS (nombre | color)",
        "RUTA COMPLETA", "DETALLE ERROR"
    ]
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border
    ws.row_dimensions[2].height = 20

    for idx, r in enumerate(resultados, 1):
        row  = idx + 2
        alt  = (row % 2 == 0)
        bg_a = "EEF4FF" if alt else None
        d    = r.get("detalle", {})
        es_error = r["estado"] == "ERROR"

        lk  = d.get("layer_k")  or {}
        lk2 = d.get("layer_k2") or {}

        lista_layers = " | ".join(
            f"{l['nombre']}({l['color_texto']})"
            for l in sorted(r.get("layers", []), key=lambda x: x["nombre"])
        )

        _c(ws, row,  1, idx,                    center=True, bg=bg_a)
        _c(ws, row,  2, r["archivo"],           bg=bg_a)
        _c(ws, row,  3, r["estado"],            center=True, bold=True,
           bg=BG.get(r["estado"], "FFFFFF"))
        _c(ws, row,  4, "SÍ" if d.get("tiene_k")  else "NO", center=True,
           bg="C6EFCE" if d.get("tiene_k")  else ("FFCCCC" if not alt else bg_a))
        _c(ws, row,  5, lk.get("nombre", ""),  bg=bg_a)
        _c(ws, row,  6, lk.get("color_texto", ""), bg=bg_a)
        _c(ws, row,  7, "SÍ ✓" if d.get("k_es_azul") else "NO ✗", center=True,
           bg="C6EFCE" if d.get("k_es_azul") else "FFCCCC")
        _c(ws, row,  8, "SÍ" if d.get("tiene_k2") else "NO", center=True,
           bg="C6EFCE" if d.get("tiene_k2") else ("FFCCCC" if not alt else bg_a))
        _c(ws, row,  9, lk2.get("nombre", ""), bg=bg_a)
        _c(ws, row, 10, lk2.get("color_texto", ""), bg=bg_a)
        _c(ws, row, 11, "SÍ ✓" if d.get("k2_es_verde") else "NO ✗", center=True,
           bg="C6EFCE" if d.get("k2_es_verde") else "FFCCCC")
        _c(ws, row, 12, len(r.get("layers", [])), center=True, bg=bg_a)
        _c(ws, row, 13, lista_layers, bg=bg_a, fuente="Courier New", size=8)
        _c(ws, row, 14, r["ruta"], bg=bg_a, color_txt="0070C0", fuente="Courier New", size=8)
        _c(ws, row, 15, r.get("detalle_error", "") if es_error else "",
           bg=BG["ERROR"] if es_error else bg_a)

    ANCHOS = [5, 35, 16, 10, 14, 16, 12, 10, 14, 16, 12, 14, 70, 80, 40]
    for i, w in enumerate(ANCHOS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    wb.save(ruta_salida)
    ok(f"Excel de prueba guardado: {ruta_salida}")


# ──────────────────────────────────────────────────────────
# SELECCIÓN INTERACTIVA DE ARCHIVOS (si lista vacía)
# ──────────────────────────────────────────────────────────
def pedir_archivos_interactivo():
    print()
    sep("═")
    print("  No hay archivos en ARCHIVOS_PRUEBA.")
    print("  Ingresa rutas DWG una por una. Enter vacío para terminar.")
    sep("═")
    archivos = []
    while True:
        ruta = input(f"  Archivo {len(archivos)+1}: ").strip().strip('"').strip("'")
        if not ruta:
            break
        if not ruta.lower().endswith(".dwg"):
            warn("No es un .dwg, se ignora")
            continue
        archivos.append(ruta)
        ok(f"Agregado: {os.path.basename(ruta)}")
    return archivos


# ──────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────
def main():
    sep("═")
    print("  TEST — AUDITORÍA LAYERS K / K2")
    print("  Script de prueba para validar lógica antes del escaneo completo")
    sep("═")

    archivos = [a for a in ARCHIVOS_PRUEBA if a.strip()]
    if not archivos:
        archivos = pedir_archivos_interactivo()

    if not archivos:
        warn("Sin archivos para probar. Agrega rutas en ARCHIVOS_PRUEBA o ingrésalas.")
        return

    print(f"\n  {len(archivos)} archivo(s) a analizar\n")

    acad = conectar_autocad()

    resultados = []
    for i, ruta in enumerate(archivos, 1):
        r = analizar_archivo(acad, ruta, i, len(archivos))
        resultados.append(r)
        time.sleep(0.3)

    try:
        pythoncom.CoUninitialize()
    except Exception:
        pass

    # Resumen en consola
    sep("═")
    print("\n  RESUMEN FINAL\n")
    for r in resultados:
        icono = {"ACTUALIZADA": "✓", "VIEJA": "✗", "INCOMPLETA": "!", "ERROR": "E"}.get(r["estado"], "?")
        print(f"  {icono}  {r['estado']:<14}  {r['archivo']}")

    tot = len(resultados)
    print(f"\n  Total      : {tot}")
    for estado in ["ACTUALIZADA", "VIEJA", "INCOMPLETA", "ERROR"]:
        n = sum(1 for r in resultados if r["estado"] == estado)
        if n:
            print(f"  {estado:<14}: {n}")

    # Guardar Excel
    nombre_excel = f"TEST_layers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print()
    generar_excel_prueba(resultados, nombre_excel)

    sep("═")
    print(f"\n  Prueba completada. Revisa el Excel: {nombre_excel}\n")
    sep("═")


if __name__ == "__main__":
    main()
